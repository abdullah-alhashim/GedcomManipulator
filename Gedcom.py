from sys import exit
from gedcom.element.individual import IndividualElement
from gedcom.parser import Parser
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import csv


class GedcomManipulator:
    def __init__(self, file_path):
        self.gedcom_parser = Parser()
        self.gedcom_parser.parse_file(file_path, False)  # Disable strict parsing
        self.root_child_elements = self.gedcom_parser.get_root_child_elements()
        self.fullName = []

    @staticmethod
    def first_name(el):
        "Returns the first name of the element (joined)"
        return ''.join(el.get_name()[0])

    def get_full_name(self, el):
        """Returns the full names of all elements in a gedcome file (.ged)"""

        self.fullName = [self.first_name(el)]
        parent = el
        while parent:  # while parent is not an empty list
            parent = self.gedcom_parser.get_parents(parent)
            if parent:
                self.fullName.append(self.first_name(parent[0]))
                parent = parent[0]

        if el.get_name()[1]:
            self.fullName.append(el.get_name()[1])

        return self.fullName

    def write_csv(self, output_file_path):
        pointer = [el.get_pointer() for el in self.root_child_elements[1:2260]]
        wb = Workbook()
        ws = wb.active
        # with open('/Users/abdullahalhashim/Desktop/Des/Family Tree/Python/FamNames.csv', mode='w') as names_files:
        #   name_writer = csv.writer(names_files, delimiter=',')
        row = 2
        for el in self.root_child_elements[1:]:
            if el.get_tag() == 'FAM':  # if element tag is "Individual," extract full name.
                for child in el.get_child_elements():
                    element = self.root_child_elements[pointer.index(child.get_value())+1]
                    full_name = self.get_full_name(element)
                    # print(full_name)
                    if child.get_tag() != 'CHIL':
                        # name_writer.writerow(full_name[::-1])
                        for col, val in enumerate(full_name[::-1], start=1):
                            cell = ws.cell(row=row, column=col+1)
                            cell.value = val
                            if child.get_tag() == 'HUSB':
                                cell.fill = PatternFill("solid", fgColor="66CCFF")
                            elif child.get_tag() == 'WIFE':
                                cell.fill = PatternFill("solid", fgColor="FFCCFF")
                        row += 1
                    else:
                        # name_writer.writerow([self.first_name(element)])
                        cell = ws.cell(row=row, column=2)
                        cell.value = self.first_name(element)
                        cell.fill = PatternFill("solid", fgColor="00CCCC")
                        row += 1
                row += 1
                # name_writer.writerow('\n')
            else:  # else terminate, i.e. if tag is "family"
                pass
        wb.save(output_file_path)


if __name__ == "__main__":
    file_path = '/path/to/gedcom/file.ged'
    output_file_path = '/path/to/output/file.xlsx'
    gedMan = GedcomManipulator(file_path)
    gedMan.write_csv(output_file_path)

